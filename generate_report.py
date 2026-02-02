"""
Enhancement Jobs Report Generator
Pulls data from Jobber API and generates monthly Excel report
Runs weekly on Fridays via GitHub Actions

Supports two modes:
1. API Mode: Fetches data directly from Jobber GraphQL API
2. CSV Mode: Reads from exported CSV files in Downloads folder

CSV fallback is used when API doesn't return expected data or for local testing.
"""

import os
import csv
import glob
import json
import requests
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule, DataBarRule

# Configuration
API_URL = "https://api.getjobber.com/api/graphql"
API_VERSION = "2023-11-15"

# Color palette
NAVY = "1B365D"
DARK_BLUE = "2E5090"
MEDIUM_BLUE = "4472C4"
LIGHT_BLUE = "D6DCE5"
GREEN_LIGHT = "C6EFCE"
RED_LIGHT = "FFC7CE"
YELLOW_LIGHT = "FFEB9C"
ENHANCEMENT_FILL = "E2EFDA"
CONTRACTED_FILL = "DDEBF7"

# Styles
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10, name='Calibri')
title_font = Font(bold=True, size=18, color=NAVY, name='Calibri')
section_font = Font(bold=True, size=12, color=NAVY, name='Calibri')
total_fill = PatternFill(start_color=MEDIUM_BLUE, end_color=MEDIUM_BLUE, fill_type="solid")
total_font = Font(bold=True, size=10, color="FFFFFF", name='Calibri')
green_fill = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
red_fill = PatternFill(start_color=RED_LIGHT, end_color=RED_LIGHT, fill_type="solid")
alt_row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
thin_border = Border(
    left=Side(style='thin', color=LIGHT_BLUE),
    right=Side(style='thin', color=LIGHT_BLUE),
    top=Side(style='thin', color=LIGHT_BLUE),
    bottom=Side(style='thin', color=LIGHT_BLUE)
)
thick_bottom = Border(bottom=Side(style='medium', color=NAVY))
currency_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
currency_format_whole = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
percent_format = '0.0%'


# CSV file locations
DOWNLOADS_DIR = r"C:\Users\daria\Downloads"
JOBS_CSV_PATTERN = "One-off jobs_Report_*.csv"
INVOICES_CSV_PATTERN = "Invoices_Report_*.csv"


def find_latest_csv(pattern, directory=DOWNLOADS_DIR):
    """Find the most recent CSV file matching pattern"""
    search_path = os.path.join(directory, pattern)
    files = glob.glob(search_path)
    if not files:
        return None
    # Sort by modification time, newest first
    return max(files, key=os.path.getmtime)


def parse_currency(value):
    """Parse currency string to float"""
    if not value:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    # Remove $ and commas, handle parentheses for negatives
    cleaned = str(value).replace('$', '').replace(',', '').strip()
    if cleaned.startswith('(') and cleaned.endswith(')'):
        cleaned = '-' + cleaned[1:-1]
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_percentage(value):
    """Parse percentage string to float"""
    if not value:
        return 0.0
    cleaned = str(value).replace('%', '').strip()
    try:
        return float(cleaned) / 100.0
    except ValueError:
        return 0.0


def parse_date(date_str):
    """Parse date from various formats"""
    if not date_str or date_str == '-':
        return None
    date_str = date_str.strip()
    # Try multiple date formats
    formats = [
        '%Y-%m-%d',           # 2026-01-31
        '%b %d, %Y',          # Jan 31, 2026
        '%B %d, %Y',          # January 31, 2026
        '%m/%d/%Y',           # 01/31/2026
        '%m/%d/%y',           # 01/31/26
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def load_jobs_from_csv(csv_path, start_date, end_date):
    """Load and filter jobs from CSV export"""
    jobs = []

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Check job type - only Enhancement and Contracted Enhancement
            job_type = row.get('Job Type', '').strip()
            if job_type not in ('Enhancement', 'Contracted Enhancement'):
                continue

            # Check date range using Scheduled start date
            date_str = row.get('Scheduled start date', '')
            job_date = parse_date(date_str)
            if job_date:
                if not (start_date <= job_date <= end_date):
                    continue

            # Parse financial data
            revenue = parse_currency(row.get('Total revenue ($)', 0))
            cost = parse_currency(row.get('Total costs ($)', 0))
            profit = parse_currency(row.get('Profit ($)', 0))

            jobs.append({
                'jobNumber': row.get('Job #', ''),
                'title': row.get('Job Type', ''),  # Use Job Type as title since that's what we're filtering on
                'client': {'name': row.get('Client name', '')},
                'jobStatus': 'closed' if row.get('Closed date') else 'active',
                'startAt': row.get('Scheduled start date', ''),
                'closedAt': row.get('Closed date', ''),
                'salesperson': row.get('Salesperson', ''),
                'assignedTo': row.get('Visits assigned to', ''),
                'invoiceNumbers': row.get('Invoice #s', ''),
                'expensesTotal': parse_currency(row.get('Expenses total ($)', 0)),
                'timeTracked': row.get('Time tracked', ''),
                'labourCost': parse_currency(row.get('Labour cost total ($)', 0)),
                'jobCosting': {
                    'totalRevenue': revenue,
                    'totalCost': cost,
                },
                'total': revenue,
                '_profit': profit,
                '_profit_pct': parse_percentage(row.get('Profit %', 0)),
                '_job_type': job_type,
                '_is_enhancement': job_type == 'Enhancement',
                '_is_contracted': job_type == 'Contracted Enhancement',
            })

    return jobs


def load_invoices_from_csv(csv_path, job_numbers):
    """Load and filter invoices from CSV export"""
    invoices = []

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Check if invoice is related to any of our job numbers
            invoice_jobs = row.get('Job #s', '')
            related = False
            for jn in job_numbers:
                if str(jn) in invoice_jobs:
                    related = True
                    break

            if not related:
                continue

            invoices.append({
                'invoiceNumber': row.get('Invoice #', ''),
                'client': {'name': row.get('Client name', '')},
                'job': {'jobNumber': invoice_jobs.split(',')[0].strip() if invoice_jobs else ''},
                'createdDate': row.get('Created date', ''),
                'issuedDate': row.get('Issued date', ''),
                'dueDate': row.get('Due date', ''),
                'lateBy': row.get('Late by', ''),
                'paidDate': row.get('Marked paid date', ''),
                'daysToPaid': row.get('Days to paid', ''),
                'lastContacted': row.get('Last contacted', ''),
                'invoiceStatus': row.get('Status', ''),
                'total': parse_currency(row.get('Total ($)', 0)),
                'balance': parse_currency(row.get('Balance ($)', 0)),
            })

    return invoices


def get_access_token():
    """Get access token from environment or refresh if needed"""
    access_token = os.environ.get('JOBBER_ACCESS_TOKEN')
    refresh_token = os.environ.get('JOBBER_REFRESH_TOKEN')
    client_id = os.environ.get('JOBBER_CLIENT_ID')
    client_secret = os.environ.get('JOBBER_CLIENT_SECRET')

    # Try existing token first
    if access_token:
        test_query = '{ jobs(first: 1) { totalCount } }'
        response = requests.post(
            API_URL,
            headers={
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json',
                'X-JOBBER-GRAPHQL-VERSION': API_VERSION
            },
            json={'query': test_query}
        )
        if response.status_code == 200 and 'errors' not in response.json():
            return access_token

    # Refresh token if available
    if refresh_token and client_id and client_secret:
        response = requests.post(
            'https://api.getjobber.com/api/oauth/token',
            data={
                'client_id': client_id,
                'client_secret': client_secret,
                'refresh_token': refresh_token,
                'grant_type': 'refresh_token'
            }
        )
        if response.status_code == 200:
            tokens = response.json()
            return tokens['access_token']

    raise Exception("Unable to get valid access token")


def fetch_jobs(access_token, start_date, end_date):
    """Fetch all jobs within date range from Jobber API"""
    all_jobs = []
    cursor = None

    while True:
        query = '''
        query($cursor: String) {
            jobs(first: 100, after: $cursor) {
                nodes {
                    jobNumber
                    title
                    jobStatus
                    startAt
                    endAt
                    total
                    client {
                        name
                    }
                    jobCosting {
                        totalRevenue
                        totalCost
                    }
                    invoices {
                        nodes {
                            invoiceNumber
                            total
                            invoiceStatus
                        }
                    }
                    visits {
                        nodes {
                            assignedUsers {
                                nodes {
                                    name {
                                        full
                                    }
                                }
                            }
                        }
                    }
                    customFields {
                        nodes {
                            label
                            valueText
                        }
                    }
                }
                pageInfo {
                    hasNextPage
                    endCursor
                }
            }
        }
        '''

        response = requests.post(
            API_URL,
            headers={
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json',
                'X-JOBBER-GRAPHQL-VERSION': API_VERSION
            },
            json={'query': query, 'variables': {'cursor': cursor}}
        )

        if response.status_code != 200:
            print(f"API Error: {response.status_code} - {response.text}")
            break

        data = response.json()
        if 'errors' in data:
            print(f"GraphQL Error: {data['errors']}")
            break

        jobs = data['data']['jobs']['nodes']

        # Filter jobs by date and type
        for job in jobs:
            job_start = job.get('startAt')
            if job_start:
                job_date = datetime.fromisoformat(job_start.replace('Z', '+00:00'))
                if start_date <= job_date.date() <= end_date:
                    # Check if it's an enhancement job (by title or custom field)
                    title_lower = (job.get('title') or '').lower()
                    is_enhancement = 'enhancement' in title_lower
                    is_contracted = 'contracted' in title_lower and 'enhancement' in title_lower

                    # Also check custom fields for job type
                    for cf in job.get('customFields', {}).get('nodes', []):
                        if cf.get('label', '').lower() == 'job type':
                            val = (cf.get('valueText') or '').lower()
                            if 'enhancement' in val:
                                is_enhancement = True
                            if 'contracted' in val:
                                is_contracted = True

                    if is_enhancement or is_contracted:
                        job['_is_contracted'] = is_contracted
                        job['_is_enhancement'] = is_enhancement and not is_contracted
                        all_jobs.append(job)

        # Pagination
        page_info = data['data']['jobs']['pageInfo']
        if not page_info['hasNextPage']:
            break
        cursor = page_info['endCursor']

        # Rate limit protection
        import time
        time.sleep(0.5)

    return all_jobs


def fetch_invoices(access_token, job_numbers):
    """Fetch invoices for specific job numbers"""
    all_invoices = []
    cursor = None

    while True:
        query = '''
        query($cursor: String) {
            invoices(first: 100, after: $cursor) {
                nodes {
                    invoiceNumber
                    total
                    invoiceStatus
                    issuedDate
                    dueDate
                    client {
                        name
                    }
                    job {
                        jobNumber
                    }
                }
                pageInfo {
                    hasNextPage
                    endCursor
                }
            }
        }
        '''

        response = requests.post(
            API_URL,
            headers={
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json',
                'X-JOBBER-GRAPHQL-VERSION': API_VERSION
            },
            json={'query': query, 'variables': {'cursor': cursor}}
        )

        if response.status_code != 200:
            break

        data = response.json()
        if 'errors' in data:
            break

        invoices = data['data']['invoices']['nodes']
        for inv in invoices:
            job = inv.get('job')
            if job and job.get('jobNumber') in job_numbers:
                all_invoices.append(inv)

        page_info = data['data']['invoices']['pageInfo']
        if not page_info['hasNextPage']:
            break
        cursor = page_info['endCursor']

        import time
        time.sleep(0.5)

    return all_invoices


def get_week_start(date_obj):
    """Get the Monday of the week for a given date"""
    if isinstance(date_obj, str):
        date_obj = parse_date(date_obj)
    if not date_obj:
        return None
    return date_obj - timedelta(days=date_obj.weekday())


def generate_report(jobs, invoices, month, year):
    """Generate comprehensive Excel report from job and invoice data"""
    from collections import defaultdict

    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Dashboard"

    month_name = datetime(year, month, 1).strftime('%B')

    # ========== CALCULATE ALL METRICS ==========

    # Basic counts
    enhancement_jobs = [j for j in jobs if j.get('_is_enhancement')]
    contracted_jobs = [j for j in jobs if j.get('_is_contracted')]

    # Financial totals
    total_revenue = sum(float(j.get('jobCosting', {}).get('totalRevenue') or j.get('total') or 0) for j in jobs)
    total_cost = sum(float(j.get('jobCosting', {}).get('totalCost') or 0) for j in jobs)
    total_profit = total_revenue - total_cost
    total_labour = sum(float(j.get('labourCost') or 0) for j in jobs)
    total_expenses = sum(float(j.get('expensesTotal') or 0) for j in jobs)
    total_materials = total_cost - total_labour  # Estimate materials as non-labor costs

    # Invoice status breakdown
    inv_paid = [i for i in invoices if i.get('invoiceStatus', '').lower() == 'paid']
    inv_awaiting = [i for i in invoices if i.get('invoiceStatus', '').lower() in ('awaiting payment', 'sent', 'viewed')]
    inv_past_due = [i for i in invoices if i.get('invoiceStatus', '').lower() == 'past due']
    inv_draft = [i for i in invoices if i.get('invoiceStatus', '').lower() == 'draft']

    paid_total = sum(float(i.get('total') or 0) for i in inv_paid)
    awaiting_total = sum(float(i.get('total') or 0) for i in inv_awaiting)
    past_due_total = sum(float(i.get('total') or 0) for i in inv_past_due)
    draft_total = sum(float(i.get('total') or 0) for i in inv_draft)
    outstanding_balance = sum(float(i.get('balance') or 0) for i in invoices)

    # Team performance
    team_stats = defaultdict(lambda: {'jobs': 0, 'revenue': 0, 'cost': 0, 'labour': 0, 'profit': 0})
    for job in jobs:
        team = job.get('assignedTo') or 'Unassigned'
        revenue = float(job.get('jobCosting', {}).get('totalRevenue') or job.get('total') or 0)
        cost = float(job.get('jobCosting', {}).get('totalCost') or 0)
        labour = float(job.get('labourCost') or 0)
        team_stats[team]['jobs'] += 1
        team_stats[team]['revenue'] += revenue
        team_stats[team]['cost'] += cost
        team_stats[team]['labour'] += labour
        team_stats[team]['profit'] += revenue - cost

    # Salesperson performance
    sales_stats = defaultdict(lambda: {'jobs': 0, 'revenue': 0, 'profit': 0})
    for job in jobs:
        salesperson = job.get('salesperson') or 'Unassigned'
        revenue = float(job.get('jobCosting', {}).get('totalRevenue') or job.get('total') or 0)
        cost = float(job.get('jobCosting', {}).get('totalCost') or 0)
        sales_stats[salesperson]['jobs'] += 1
        sales_stats[salesperson]['revenue'] += revenue
        sales_stats[salesperson]['profit'] += revenue - cost

    # Weekly breakdown
    weekly_stats = defaultdict(lambda: {'jobs': 0, 'enhancement': 0, 'contracted': 0, 'revenue': 0, 'cost': 0, 'profit': 0})
    for job in jobs:
        week_start = get_week_start(job.get('startAt'))
        if week_start:
            week_key = week_start.strftime('%Y-%m-%d')
            revenue = float(job.get('jobCosting', {}).get('totalRevenue') or job.get('total') or 0)
            cost = float(job.get('jobCosting', {}).get('totalCost') or 0)
            weekly_stats[week_key]['jobs'] += 1
            weekly_stats[week_key]['enhancement'] += 1 if job.get('_is_enhancement') else 0
            weekly_stats[week_key]['contracted'] += 1 if job.get('_is_contracted') else 0
            weekly_stats[week_key]['revenue'] += revenue
            weekly_stats[week_key]['cost'] += cost
            weekly_stats[week_key]['profit'] += revenue - cost

    # Client analysis
    client_stats = defaultdict(lambda: {'jobs': 0, 'revenue': 0, 'cost': 0, 'profit': 0})
    for job in jobs:
        client = job.get('client', {}).get('name') or 'Unknown'
        revenue = float(job.get('jobCosting', {}).get('totalRevenue') or job.get('total') or 0)
        cost = float(job.get('jobCosting', {}).get('totalCost') or 0)
        client_stats[client]['jobs'] += 1
        client_stats[client]['revenue'] += revenue
        client_stats[client]['cost'] += cost
        client_stats[client]['profit'] += revenue - cost

    # ========== SHEET 1: EXECUTIVE DASHBOARD ==========

    # Set column widths
    for col, width in {'A': 3, 'B': 20, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15, 'I': 3}.items():
        ws.column_dimensions[col].width = width

    # Title
    ws.merge_cells('B2:H2')
    ws['B2'] = f"ENHANCEMENT JOBS - {month_name.upper()} {year}"
    ws['B2'].font = title_font

    ws.merge_cells('B3:H3')
    ws['B3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['B3'].font = Font(size=10, italic=True, color="808080")

    # Row 5: KPI Headers
    row = 5

    # KPI Cards Row 1
    kpi_data = [
        ('B', 'TOTAL JOBS', len(jobs), f"{len(enhancement_jobs)} std / {len(contracted_jobs)} contracted"),
        ('D', 'TOTAL REVENUE', total_revenue, None),
        ('F', 'GROSS PROFIT', total_profit, f"margin: {(total_profit/total_revenue*100 if total_revenue else 0):.1f}%"),
        ('H', 'AVG JOB VALUE', total_revenue / len(jobs) if jobs else 0, None),
    ]

    for col_letter, label, value, subtext in kpi_data:
        ws[f'{col_letter}{row}'] = label
        ws[f'{col_letter}{row}'].font = Font(size=9, color="808080")
        ws[f'{col_letter}{row+1}'] = value
        ws[f'{col_letter}{row+1}'].font = Font(bold=True, size=20, color=NAVY)
        if isinstance(value, float):
            ws[f'{col_letter}{row+1}'].number_format = currency_format_whole
        if subtext:
            ws[f'{col_letter}{row+2}'] = subtext
            ws[f'{col_letter}{row+2}'].font = Font(size=8, color="548235" if "margin" in subtext else "808080")

    # Apply KPI backgrounds
    for col in range(2, 9):
        for r in range(row, row + 3):
            ws.cell(row=r, column=col).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    # Row 10: Cost Breakdown Section
    row = 10
    ws[f'B{row}'] = "COST BREAKDOWN"
    ws[f'B{row}'].font = section_font

    row += 1
    cost_headers = ['Category', 'Amount', '% of Total', '% of Revenue']
    for idx, h in enumerate(cost_headers):
        cell = ws.cell(row=row, column=2 + idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    cost_data = [
        ('Labor Costs', total_labour),
        ('Materials/Expenses', total_expenses),
        ('Other Costs', max(0, total_cost - total_labour - total_expenses)),
    ]

    for i, (cat, amt) in enumerate(cost_data):
        r = row + 1 + i
        ws.cell(row=r, column=2).value = cat
        ws.cell(row=r, column=3).value = amt
        ws.cell(row=r, column=3).number_format = currency_format
        ws.cell(row=r, column=4).value = amt / total_cost if total_cost else 0
        ws.cell(row=r, column=4).number_format = percent_format
        ws.cell(row=r, column=5).value = amt / total_revenue if total_revenue else 0
        ws.cell(row=r, column=5).number_format = percent_format
        for c in range(2, 6):
            ws.cell(row=r, column=c).border = thin_border

    # Total row
    r = row + len(cost_data) + 1
    ws.cell(row=r, column=2).value = "TOTAL COSTS"
    ws.cell(row=r, column=2).font = total_font
    ws.cell(row=r, column=2).fill = total_fill
    ws.cell(row=r, column=3).value = total_cost
    ws.cell(row=r, column=3).number_format = currency_format
    ws.cell(row=r, column=3).font = total_font
    ws.cell(row=r, column=3).fill = total_fill
    ws.cell(row=r, column=4).value = 1.0
    ws.cell(row=r, column=4).number_format = percent_format
    ws.cell(row=r, column=4).font = total_font
    ws.cell(row=r, column=4).fill = total_fill
    ws.cell(row=r, column=5).value = total_cost / total_revenue if total_revenue else 0
    ws.cell(row=r, column=5).number_format = percent_format
    ws.cell(row=r, column=5).font = total_font
    ws.cell(row=r, column=5).fill = total_fill

    # Row 18: Invoice Status Section
    row = 18
    ws[f'B{row}'] = "INVOICE STATUS"
    ws[f'B{row}'].font = section_font

    row += 1
    inv_headers = ['Status', 'Count', 'Amount', '% of Total', 'Avg Invoice']
    for idx, h in enumerate(inv_headers):
        cell = ws.cell(row=row, column=2 + idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    inv_status_data = [
        ('Paid', len(inv_paid), paid_total, GREEN_LIGHT),
        ('Awaiting Payment', len(inv_awaiting), awaiting_total, YELLOW_LIGHT),
        ('Past Due', len(inv_past_due), past_due_total, RED_LIGHT),
        ('Draft', len(inv_draft), draft_total, LIGHT_BLUE),
    ]

    total_inv_amt = paid_total + awaiting_total + past_due_total + draft_total
    for i, (status, count, amt, fill_color) in enumerate(inv_status_data):
        r = row + 1 + i
        ws.cell(row=r, column=2).value = status
        ws.cell(row=r, column=2).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws.cell(row=r, column=3).value = count
        ws.cell(row=r, column=4).value = amt
        ws.cell(row=r, column=4).number_format = currency_format
        ws.cell(row=r, column=5).value = amt / total_inv_amt if total_inv_amt else 0
        ws.cell(row=r, column=5).number_format = percent_format
        ws.cell(row=r, column=6).value = amt / count if count else 0
        ws.cell(row=r, column=6).number_format = currency_format
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = thin_border

    # Outstanding balance
    r = row + len(inv_status_data) + 2
    ws.cell(row=r, column=2).value = "Outstanding Balance:"
    ws.cell(row=r, column=2).font = Font(bold=True, color=NAVY)
    ws.cell(row=r, column=4).value = outstanding_balance
    ws.cell(row=r, column=4).number_format = currency_format
    ws.cell(row=r, column=4).font = Font(bold=True, color="C00000" if outstanding_balance > 0 else "548235")

    # Row 27: Job Type Comparison
    row = 27
    ws[f'B{row}'] = "ENHANCEMENT TYPE COMPARISON"
    ws[f'B{row}'].font = section_font

    row += 1
    type_headers = ['Job Type', 'Jobs', 'Revenue', 'Costs', 'Profit', 'Margin %', 'Avg Value']
    for idx, h in enumerate(type_headers):
        cell = ws.cell(row=row, column=2 + idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    type_data = [
        ('Enhancement', enhancement_jobs, ENHANCEMENT_FILL),
        ('Contracted Enhancement', contracted_jobs, CONTRACTED_FILL),
    ]

    for i, (jtype, jlist, fill_color) in enumerate(type_data):
        r = row + 1 + i
        rev = sum(float(j.get('jobCosting', {}).get('totalRevenue') or j.get('total') or 0) for j in jlist)
        cost = sum(float(j.get('jobCosting', {}).get('totalCost') or 0) for j in jlist)
        profit = rev - cost
        ws.cell(row=r, column=2).value = jtype
        ws.cell(row=r, column=2).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws.cell(row=r, column=3).value = len(jlist)
        ws.cell(row=r, column=4).value = rev
        ws.cell(row=r, column=4).number_format = currency_format
        ws.cell(row=r, column=5).value = cost
        ws.cell(row=r, column=5).number_format = currency_format
        ws.cell(row=r, column=6).value = profit
        ws.cell(row=r, column=6).number_format = currency_format
        ws.cell(row=r, column=7).value = profit / rev if rev else 0
        ws.cell(row=r, column=7).number_format = percent_format
        ws.cell(row=r, column=8).value = rev / len(jlist) if jlist else 0
        ws.cell(row=r, column=8).number_format = currency_format
        for c in range(2, 9):
            ws.cell(row=r, column=c).border = thin_border

    # Total row
    r = row + 3
    ws.cell(row=r, column=2).value = "TOTAL"
    ws.cell(row=r, column=2).font = total_font
    ws.cell(row=r, column=2).fill = total_fill
    for c, val in [(3, len(jobs)), (4, total_revenue), (5, total_cost), (6, total_profit)]:
        ws.cell(row=r, column=c).value = val
        ws.cell(row=r, column=c).font = total_font
        ws.cell(row=r, column=c).fill = total_fill
        if c >= 4:
            ws.cell(row=r, column=c).number_format = currency_format
    ws.cell(row=r, column=7).value = total_profit / total_revenue if total_revenue else 0
    ws.cell(row=r, column=7).number_format = percent_format
    ws.cell(row=r, column=7).font = total_font
    ws.cell(row=r, column=7).fill = total_fill
    ws.cell(row=r, column=8).value = total_revenue / len(jobs) if jobs else 0
    ws.cell(row=r, column=8).number_format = currency_format
    ws.cell(row=r, column=8).font = total_font
    ws.cell(row=r, column=8).fill = total_fill

    # ========== SHEET 2: WEEKLY TRENDS ==========

    ws_weekly = wb.create_sheet("Weekly Trends")

    ws_weekly['B2'] = f"WEEKLY PERFORMANCE - {month_name.upper()} {year}"
    ws_weekly['B2'].font = title_font

    headers = ['Week Starting', 'Enhancement', 'Contracted', 'Total Jobs', 'Revenue', 'Costs', 'Profit', 'Margin %', 'Rev WoW %']
    for idx, h in enumerate(headers):
        cell = ws_weekly.cell(row=4, column=2 + idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    sorted_weeks = sorted(weekly_stats.keys())
    prev_revenue = None
    for i, week in enumerate(sorted_weeks):
        r = 5 + i
        stats = weekly_stats[week]
        margin = stats['profit'] / stats['revenue'] if stats['revenue'] else 0
        wow_change = (stats['revenue'] - prev_revenue) / prev_revenue if prev_revenue and prev_revenue != 0 else 0

        ws_weekly.cell(row=r, column=2).value = week
        ws_weekly.cell(row=r, column=3).value = stats['enhancement']
        ws_weekly.cell(row=r, column=3).fill = PatternFill(start_color=ENHANCEMENT_FILL, end_color=ENHANCEMENT_FILL, fill_type="solid")
        ws_weekly.cell(row=r, column=4).value = stats['contracted']
        ws_weekly.cell(row=r, column=4).fill = PatternFill(start_color=CONTRACTED_FILL, end_color=CONTRACTED_FILL, fill_type="solid")
        ws_weekly.cell(row=r, column=5).value = stats['jobs']
        ws_weekly.cell(row=r, column=6).value = stats['revenue']
        ws_weekly.cell(row=r, column=6).number_format = currency_format
        ws_weekly.cell(row=r, column=7).value = stats['cost']
        ws_weekly.cell(row=r, column=7).number_format = currency_format
        ws_weekly.cell(row=r, column=8).value = stats['profit']
        ws_weekly.cell(row=r, column=8).number_format = currency_format
        ws_weekly.cell(row=r, column=8).fill = green_fill if stats['profit'] > 0 else red_fill
        ws_weekly.cell(row=r, column=9).value = margin
        ws_weekly.cell(row=r, column=9).number_format = percent_format
        ws_weekly.cell(row=r, column=10).value = wow_change if prev_revenue else None
        ws_weekly.cell(row=r, column=10).number_format = percent_format

        for c in range(2, 11):
            ws_weekly.cell(row=r, column=c).border = thin_border
            if i % 2 == 1:
                if c not in [3, 4, 8]:  # Don't override special fills
                    ws_weekly.cell(row=r, column=c).fill = alt_row_fill

        prev_revenue = stats['revenue']

    for col, width in {'B': 14, 'C': 12, 'D': 12, 'E': 12, 'F': 14, 'G': 14, 'H': 14, 'I': 12, 'J': 12}.items():
        ws_weekly.column_dimensions[col].width = width

    # ========== SHEET 3: TEAM PERFORMANCE ==========

    ws_team = wb.create_sheet("Team Performance")

    ws_team['B2'] = f"TEAM P&L - {month_name.upper()} {year}"
    ws_team['B2'].font = title_font

    headers = ['Team/Crew', 'Jobs', 'Revenue', 'Labor Cost', 'Total Cost', 'Profit', 'Margin %', 'Avg Job Value']
    for idx, h in enumerate(headers):
        cell = ws_team.cell(row=4, column=2 + idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    sorted_teams = sorted(team_stats.items(), key=lambda x: x[1]['revenue'], reverse=True)
    for i, (team, stats) in enumerate(sorted_teams):
        r = 5 + i
        margin = stats['profit'] / stats['revenue'] if stats['revenue'] else 0

        ws_team.cell(row=r, column=2).value = team
        ws_team.cell(row=r, column=3).value = stats['jobs']
        ws_team.cell(row=r, column=4).value = stats['revenue']
        ws_team.cell(row=r, column=4).number_format = currency_format
        ws_team.cell(row=r, column=5).value = stats['labour']
        ws_team.cell(row=r, column=5).number_format = currency_format
        ws_team.cell(row=r, column=6).value = stats['cost']
        ws_team.cell(row=r, column=6).number_format = currency_format
        ws_team.cell(row=r, column=7).value = stats['profit']
        ws_team.cell(row=r, column=7).number_format = currency_format
        ws_team.cell(row=r, column=7).fill = green_fill if stats['profit'] > 0 else red_fill
        ws_team.cell(row=r, column=8).value = margin
        ws_team.cell(row=r, column=8).number_format = percent_format
        ws_team.cell(row=r, column=9).value = stats['revenue'] / stats['jobs'] if stats['jobs'] else 0
        ws_team.cell(row=r, column=9).number_format = currency_format

        for c in range(2, 10):
            ws_team.cell(row=r, column=c).border = thin_border
            if i % 2 == 1 and c != 7:
                ws_team.cell(row=r, column=c).fill = alt_row_fill

    for col, width in {'B': 18, 'C': 8, 'D': 14, 'E': 14, 'F': 14, 'G': 14, 'H': 12, 'I': 14}.items():
        ws_team.column_dimensions[col].width = width

    # ========== SHEET 4: SALESPERSON PERFORMANCE ==========

    ws_sales = wb.create_sheet("Salesperson")

    ws_sales['B2'] = f"SALESPERSON PERFORMANCE - {month_name.upper()} {year}"
    ws_sales['B2'].font = title_font

    headers = ['Salesperson', 'Jobs', 'Revenue', 'Profit', 'Margin %', 'Avg Job Value']
    for idx, h in enumerate(headers):
        cell = ws_sales.cell(row=4, column=2 + idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    sorted_sales = sorted(sales_stats.items(), key=lambda x: x[1]['revenue'], reverse=True)
    for i, (salesperson, stats) in enumerate(sorted_sales):
        r = 5 + i
        margin = stats['profit'] / stats['revenue'] if stats['revenue'] else 0

        ws_sales.cell(row=r, column=2).value = salesperson
        ws_sales.cell(row=r, column=3).value = stats['jobs']
        ws_sales.cell(row=r, column=4).value = stats['revenue']
        ws_sales.cell(row=r, column=4).number_format = currency_format
        ws_sales.cell(row=r, column=5).value = stats['profit']
        ws_sales.cell(row=r, column=5).number_format = currency_format
        ws_sales.cell(row=r, column=5).fill = green_fill if stats['profit'] > 0 else red_fill
        ws_sales.cell(row=r, column=6).value = margin
        ws_sales.cell(row=r, column=6).number_format = percent_format
        ws_sales.cell(row=r, column=7).value = stats['revenue'] / stats['jobs'] if stats['jobs'] else 0
        ws_sales.cell(row=r, column=7).number_format = currency_format

        for c in range(2, 8):
            ws_sales.cell(row=r, column=c).border = thin_border
            if i % 2 == 1 and c != 5:
                ws_sales.cell(row=r, column=c).fill = alt_row_fill

    for col, width in {'B': 20, 'C': 8, 'D': 14, 'E': 14, 'F': 12, 'G': 14}.items():
        ws_sales.column_dimensions[col].width = width

    # ========== SHEET 5: CLIENT ANALYSIS ==========

    ws_client = wb.create_sheet("Client Analysis")

    ws_client['B2'] = f"TOP CLIENTS - {month_name.upper()} {year}"
    ws_client['B2'].font = title_font

    headers = ['Client', 'Jobs', 'Revenue', 'Costs', 'Profit', 'Margin %', 'Avg Job Value']
    for idx, h in enumerate(headers):
        cell = ws_client.cell(row=4, column=2 + idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    sorted_clients = sorted(client_stats.items(), key=lambda x: x[1]['revenue'], reverse=True)[:20]  # Top 20
    for i, (client, stats) in enumerate(sorted_clients):
        r = 5 + i
        margin = stats['profit'] / stats['revenue'] if stats['revenue'] else 0

        ws_client.cell(row=r, column=2).value = client
        ws_client.cell(row=r, column=3).value = stats['jobs']
        ws_client.cell(row=r, column=4).value = stats['revenue']
        ws_client.cell(row=r, column=4).number_format = currency_format
        ws_client.cell(row=r, column=5).value = stats['cost']
        ws_client.cell(row=r, column=5).number_format = currency_format
        ws_client.cell(row=r, column=6).value = stats['profit']
        ws_client.cell(row=r, column=6).number_format = currency_format
        ws_client.cell(row=r, column=6).fill = green_fill if stats['profit'] > 0 else red_fill
        ws_client.cell(row=r, column=7).value = margin
        ws_client.cell(row=r, column=7).number_format = percent_format
        ws_client.cell(row=r, column=8).value = stats['revenue'] / stats['jobs'] if stats['jobs'] else 0
        ws_client.cell(row=r, column=8).number_format = currency_format

        for c in range(2, 9):
            ws_client.cell(row=r, column=c).border = thin_border
            if i % 2 == 1 and c != 6:
                ws_client.cell(row=r, column=c).fill = alt_row_fill

    for col, width in {'B': 25, 'C': 8, 'D': 14, 'E': 14, 'F': 14, 'G': 12, 'H': 14}.items():
        ws_client.column_dimensions[col].width = width

    # ========== SHEET 6: JOBS (RAW DATA) ==========

    ws_jobs = wb.create_sheet("Jobs")
    headers = ['Job #', 'Client', 'Type', 'Status', 'Start Date', 'Closed Date', 'Salesperson', 'Team',
               'Revenue', 'Labor Cost', 'Expenses', 'Total Cost', 'Profit', 'Profit %']
    for col, header in enumerate(headers, 1):
        cell = ws_jobs.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, job in enumerate(jobs, 2):
        costing = job.get('jobCosting') or {}
        revenue = float(costing.get('totalRevenue') or job.get('total') or 0)
        cost = float(costing.get('totalCost') or 0)
        profit = revenue - cost
        job_type = 'Contracted Enhancement' if job.get('_is_contracted') else 'Enhancement'
        labour = float(job.get('labourCost') or 0)
        expenses = float(job.get('expensesTotal') or 0)

        ws_jobs.cell(row=row_idx, column=1).value = job.get('jobNumber')
        ws_jobs.cell(row=row_idx, column=2).value = job.get('client', {}).get('name', '')
        ws_jobs.cell(row=row_idx, column=3).value = job_type
        ws_jobs.cell(row=row_idx, column=3).fill = PatternFill(
            start_color=CONTRACTED_FILL if job.get('_is_contracted') else ENHANCEMENT_FILL,
            end_color=CONTRACTED_FILL if job.get('_is_contracted') else ENHANCEMENT_FILL,
            fill_type="solid"
        )
        ws_jobs.cell(row=row_idx, column=4).value = job.get('jobStatus', '')
        ws_jobs.cell(row=row_idx, column=5).value = job.get('startAt', '')
        ws_jobs.cell(row=row_idx, column=6).value = job.get('closedAt', '')
        ws_jobs.cell(row=row_idx, column=7).value = job.get('salesperson', '')
        ws_jobs.cell(row=row_idx, column=8).value = job.get('assignedTo', '')
        ws_jobs.cell(row=row_idx, column=9).value = revenue
        ws_jobs.cell(row=row_idx, column=9).number_format = currency_format
        ws_jobs.cell(row=row_idx, column=10).value = labour
        ws_jobs.cell(row=row_idx, column=10).number_format = currency_format
        ws_jobs.cell(row=row_idx, column=11).value = expenses
        ws_jobs.cell(row=row_idx, column=11).number_format = currency_format
        ws_jobs.cell(row=row_idx, column=12).value = cost
        ws_jobs.cell(row=row_idx, column=12).number_format = currency_format
        ws_jobs.cell(row=row_idx, column=13).value = profit
        ws_jobs.cell(row=row_idx, column=13).number_format = currency_format
        ws_jobs.cell(row=row_idx, column=14).value = profit / revenue if revenue else 0
        ws_jobs.cell(row=row_idx, column=14).number_format = percent_format

        for col in range(1, 15):
            ws_jobs.cell(row=row_idx, column=col).border = thin_border
            if row_idx % 2 == 0 and col != 3:
                ws_jobs.cell(row=row_idx, column=col).fill = alt_row_fill

    # Conditional formatting for profit
    if len(jobs) > 0:
        ws_jobs.conditional_formatting.add(f'M2:M{len(jobs)+1}', FormulaRule(formula=['$M2>0'], fill=green_fill))
        ws_jobs.conditional_formatting.add(f'M2:M{len(jobs)+1}', FormulaRule(formula=['$M2<0'], fill=red_fill))

    ws_jobs.freeze_panes = 'A2'

    for col, width in {'A': 10, 'B': 22, 'C': 20, 'D': 10, 'E': 14, 'F': 14, 'G': 16, 'H': 14,
                       'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 10}.items():
        ws_jobs.column_dimensions[col].width = width

    # ========== SHEET 7: INVOICES (RAW DATA) ==========

    ws_inv = wb.create_sheet("Invoices")
    inv_headers = ['Invoice #', 'Client', 'Job #', 'Total', 'Balance', 'Status', 'Issued', 'Due', 'Paid Date', 'Days to Paid']
    for col, header in enumerate(inv_headers, 1):
        cell = ws_inv.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, inv in enumerate(invoices, 2):
        status = inv.get('invoiceStatus', '')
        status_lower = status.lower()

        ws_inv.cell(row=row_idx, column=1).value = inv.get('invoiceNumber')
        ws_inv.cell(row=row_idx, column=2).value = inv.get('client', {}).get('name', '')
        ws_inv.cell(row=row_idx, column=3).value = inv.get('job', {}).get('jobNumber', '')
        ws_inv.cell(row=row_idx, column=4).value = float(inv.get('total') or 0)
        ws_inv.cell(row=row_idx, column=4).number_format = currency_format
        ws_inv.cell(row=row_idx, column=5).value = float(inv.get('balance') or 0)
        ws_inv.cell(row=row_idx, column=5).number_format = currency_format
        ws_inv.cell(row=row_idx, column=6).value = status

        # Color code status
        if status_lower == 'paid':
            ws_inv.cell(row=row_idx, column=6).fill = green_fill
        elif status_lower in ('awaiting payment', 'sent', 'viewed'):
            ws_inv.cell(row=row_idx, column=6).fill = PatternFill(start_color=YELLOW_LIGHT, end_color=YELLOW_LIGHT, fill_type="solid")
        elif status_lower == 'past due':
            ws_inv.cell(row=row_idx, column=6).fill = red_fill

        ws_inv.cell(row=row_idx, column=7).value = inv.get('issuedDate', '')
        ws_inv.cell(row=row_idx, column=8).value = inv.get('dueDate', '')
        ws_inv.cell(row=row_idx, column=9).value = inv.get('paidDate', '')
        ws_inv.cell(row=row_idx, column=10).value = inv.get('daysToPaid', '')

        for col in range(1, 11):
            ws_inv.cell(row=row_idx, column=col).border = thin_border

    # Highlight outstanding balances
    if len(invoices) > 0:
        ws_inv.conditional_formatting.add(f'E2:E{len(invoices)+1}', FormulaRule(formula=['$E2>0'], fill=red_fill))

    ws_inv.freeze_panes = 'A2'

    for col, width in {'A': 12, 'B': 22, 'C': 10, 'D': 12, 'E': 12, 'F': 18, 'G': 14, 'H': 14, 'I': 14, 'J': 12}.items():
        ws_inv.column_dimensions[col].width = width

    return wb


def main():
    print("Enhancement Report Generator")
    print("=" * 50)

    # Check for month override from environment or command line
    import sys
    month_override = os.environ.get('REPORT_MONTH') or (sys.argv[1] if len(sys.argv) > 1 else None)
    use_csv = os.environ.get('USE_CSV', '').lower() == 'true' or '--csv' in sys.argv

    # Get current date info
    today = datetime.now()

    # Determine report month
    if month_override:
        report_month = int(month_override)
        report_year = today.year
        # If requesting a future month, assume previous year
        if report_month > today.month:
            report_year -= 1
    else:
        report_month = today.month
        report_year = today.year

    # Date range for target month
    start_date = datetime(report_year, report_month, 1).date()
    if report_month == 12:
        end_date = datetime(report_year + 1, 1, 1).date() - timedelta(days=1)
    else:
        end_date = datetime(report_year, report_month + 1, 1).date() - timedelta(days=1)

    print(f"Report Period: {start_date} to {end_date}")

    jobs = []
    invoices = []

    # Try API first (if not forcing CSV mode)
    if not use_csv:
        try:
            print("Attempting Jobber API...")
            access_token = get_access_token()
            print("Authentication successful!")

            print("Fetching enhancement jobs from API...")
            jobs = fetch_jobs(access_token, start_date, end_date)
            print(f"Found {len(jobs)} enhancement jobs from API")

            if jobs:
                job_numbers = {j.get('jobNumber') for j in jobs if j.get('jobNumber')}
                print("Fetching related invoices from API...")
                invoices = fetch_invoices(access_token, job_numbers)
                print(f"Found {len(invoices)} related invoices")

        except Exception as e:
            print(f"API access failed: {e}")
            print("Falling back to CSV mode...")
            jobs = []

    # Use CSV fallback if API returned no data or CSV mode is forced
    if not jobs:
        print("\nUsing CSV fallback mode...")

        # Find latest CSV files
        jobs_csv = find_latest_csv(JOBS_CSV_PATTERN)
        invoices_csv = find_latest_csv(INVOICES_CSV_PATTERN)

        if not jobs_csv:
            print(f"ERROR: No jobs CSV found matching pattern: {JOBS_CSV_PATTERN}")
            print(f"Please export 'One-off jobs' report from Jobber to: {DOWNLOADS_DIR}")
            return None

        print(f"Loading jobs from: {os.path.basename(jobs_csv)}")
        jobs = load_jobs_from_csv(jobs_csv, start_date, end_date)
        print(f"Found {len(jobs)} enhancement jobs in date range")

        if jobs and invoices_csv:
            job_numbers = {j.get('jobNumber') for j in jobs if j.get('jobNumber')}
            print(f"Loading invoices from: {os.path.basename(invoices_csv)}")
            invoices = load_invoices_from_csv(invoices_csv, job_numbers)
            print(f"Found {len(invoices)} related invoices")

    if not jobs:
        print("\nNo enhancement jobs found for this period.")
        print("Please ensure:")
        print("  1. Jobs exist in Jobber with Job Type = 'Enhancement' or 'Contracted Enhancement'")
        print("  2. Jobs have Scheduled start dates within the report period")
        print("  3. CSV exports are up-to-date in Downloads folder")
        return None

    # Generate report
    print("\nGenerating Excel report...")
    wb = generate_report(jobs, invoices, report_month, report_year)

    # Save report (xlsx format - openpyxl doesn't support xlsm)
    month_name = datetime(report_year, report_month, 1).strftime('%B')
    filename = f"{report_month}-OneOffReport-{month_name}.xlsx"
    wb.save(filename)
    print(f"Report saved: {filename}")

    # Print summary
    enhancement_count = sum(1 for j in jobs if j.get('_is_enhancement'))
    contracted_count = sum(1 for j in jobs if j.get('_is_contracted'))
    total_revenue = sum(float(j.get('jobCosting', {}).get('totalRevenue') or j.get('total') or 0) for j in jobs)

    print(f"\n{'='*50}")
    print(f"REPORT SUMMARY - {month_name} {report_year}")
    print(f"{'='*50}")
    print(f"Enhancement Jobs: {enhancement_count}")
    print(f"Contracted Enhancement Jobs: {contracted_count}")
    print(f"Total Jobs: {len(jobs)}")
    print(f"Total Revenue: ${total_revenue:,.2f}")
    print(f"Related Invoices: {len(invoices)}")

    return filename


if __name__ == "__main__":
    main()
